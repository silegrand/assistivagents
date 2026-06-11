"""
fetch_corridor_care.py — Assistiv Systems NHS Pressure Intelligence
Corridor Care Weekly Fetcher

Runs in GitHub Actions every Thursday at 10:00 UTC (UEC SitRep publishes at 09:30).
Scrapes the NHS England UEC SitRep 2025-26 page, finds the current weekly Excel file,
downloads it, extracts corridor care and related pressure metrics for Kent trusts,
and writes kent-corridor-data.json to the repo.

Kent Trusts covered:
  RVV — East Kent Hospitals University NHS Foundation Trust
        (Thanet, Dover, Folkestone & Hythe, Canterbury, Swale, Shepway)
  RWF — Maidstone and Tunbridge Wells NHS Trust
        (Maidstone, Tonbridge & Malling, Tunbridge Wells, Sevenoaks, Ashford,
         Gravesham, Dartford)

Data source: NHS England UEC Daily Situation Reports
URL: https://www.england.nhs.uk/statistics/statistical-work-areas/uec-sitrep/
Licence: Open Government Licence v3.0
"""

import os
import re
import json
import base64
import requests
from datetime import datetime, timezone
from bs4 import BeautifulSoup
import openpyxl
from io import BytesIO

# ── CONFIG ────────────────────────────────────────────────────────────
GITHUB_REPO   = "silegrand/assistiv_cloud"
GITHUB_FILE   = "kent-corridor-data.json"
GITHUB_TOKEN  = os.environ.get("GITHUB_TOKEN", "")

SITREP_INDEX  = "https://www.england.nhs.uk/statistics/statistical-work-areas/uec-sitrep/urgent-and-emergency-care-daily-situation-reports-2025-26/"
RAW_URL       = f"https://raw.githubusercontent.com/{GITHUB_REPO}/main/{GITHUB_FILE}"

# Kent trust codes (ODS codes used in NHS SitRep Excel)
KENT_TRUSTS = {
    "RVV": {
        "name": "East Kent Hospitals University NHS Foundation Trust",
        "short": "East Kent Hospitals",
        "districts": ["Thanet", "Dover", "Folkestone & Hythe", "Canterbury", "Swale"],
    },
    "RWF": {
        "name": "Maidstone and Tunbridge Wells NHS Trust",
        "short": "Maidstone & Tunbridge Wells",
        "districts": ["Maidstone", "Tonbridge & Malling", "Tunbridge Wells",
                      "Sevenoaks", "Ashford", "Gravesham", "Dartford"],
    },
}

HEADERS = {
    "User-Agent": "AssistivSystems/1.0 (NHS open data; simon@assistiv.co)"
}


def find_sitrep_excel_url(tab_name="UEC Daily SitRep"):
    """
    Scrape the UEC SitRep index page and find the current weekly Excel download URL.
    Looks for the 'Web-File-Timeseries-UEC-Daily-SitRep' link.
    """
    print(f"Scraping SitRep index: {SITREP_INDEX}")
    r = requests.get(SITREP_INDEX, headers=HEADERS, timeout=20)
    r.raise_for_status()
    soup = BeautifulSoup(r.text, "html.parser")
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if "Timeseries-UEC-Daily-SitRep" in href and href.endswith(".xlsx"):
            url = href if href.startswith("http") else "https://www.england.nhs.uk" + href
            print(f"  Found SitRep Excel: {url}")
            return url
    raise ValueError("Could not find UEC Daily SitRep Excel URL on index page")


def download_excel(url):
    """Download Excel file, return as BytesIO."""
    print(f"Downloading: {url}")
    r = requests.get(url, headers=HEADERS, timeout=60)
    r.raise_for_status()
    print(f"  Downloaded {len(r.content):,} bytes")
    return BytesIO(r.content)


def find_trust_rows(ws, trust_codes):
    """
    Scan worksheet for rows matching Kent trust ODS codes.
    Returns dict: {trust_code: row_index}
    SitRep Excel typically has trust codes in column A or B.
    """
    trust_rows = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row), start=1):
        for cell in row[:4]:  # check first 4 cols
            if cell.value and str(cell.value).strip().upper() in trust_codes:
                code = str(cell.value).strip().upper()
                if code not in trust_rows:
                    trust_rows[code] = row_idx
                    print(f"  Found {code} at row {row_idx}, col {cell.column}")
    return trust_rows


def find_column_header(ws, header_patterns, search_rows=10):
    """
    Find column indices for corridor care fields by matching header text patterns.
    Returns dict: {pattern_key: col_index}
    """
    cols = {}
    for row in ws.iter_rows(min_row=1, max_row=search_rows):
        for cell in row:
            if not cell.value:
                continue
            val = str(cell.value).lower().strip()
            for key, patterns in header_patterns.items():
                if key not in cols:
                    if any(p.lower() in val for p in patterns):
                        cols[key] = cell.column
                        print(f"  Column '{key}' → col {cell.column}: '{cell.value}'")
    return cols


def extract_trust_metrics(wb, trust_codes):
    """
    Extract corridor care and pressure metrics per trust.
    Tries each sheet in the workbook to find trust data.
    
    Corridor care fields (new from March 2026, per NHS England definition):
      - corridor_care_ed: patients in ED corridor care >45 mins
      - corridor_care_ward: patients in ward corridor care >45 mins
      - corridor_care_total: sum of above
    
    Supporting pressure metrics:
      - twelve_hour_waits: patients waiting >12h for admission decision
      - beds_occupied_pct: general & acute bed occupancy %
      - delayed_discharges: patients not meeting criteria to reside
    """
    results = {}

    # Header patterns to search for — corridor care fields added March 2026
    HEADER_PATTERNS = {
        "corridor_ed":      ["corridor", "ed corridor", "corridor care ed",
                             "care ed", "inappropriate area ed"],
        "corridor_ward":    ["corridor ward", "ward corridor", "corridor care ward",
                             "inappropriate area ward", "care ward"],
        "twelve_hour":      ["12 hour", "12-hour", "12hr", "trolley waits",
                             "decision to admit", ">12"],
        "beds_occ":         ["occupancy", "occupied beds", "beds occupied",
                             "g&a occupied", "general and acute occupied"],
        "delayed_discharge":["delayed", "no criteria", "medically fit",
                             "criteria to reside"],
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row < 5:
            continue
        print(f"\n  Scanning sheet: '{sheet_name}' ({ws.max_row} rows × {ws.max_column} cols)")

        trust_rows = find_trust_rows(ws, set(trust_codes.keys()))
        if not trust_rows:
            continue

        cols = find_column_header(ws, HEADER_PATTERNS)
        if not cols:
            print("  No matching column headers found in this sheet")
            continue

        # Find the most recent data column (rightmost non-empty numeric column
        # to the right of the header area — SitRep Excel is time-series wide)
        # Strategy: for a known trust row, scan right from col 5 for the last
        # non-None numeric value
        for code, row_idx in trust_rows.items():
            trust_data = {}
            row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx,
                                    values_only=True))[0]

            for metric_key, col_idx in cols.items():
                try:
                    val = ws.cell(row=row_idx, column=col_idx).value
                    # For time-series sheets the column found may be a header row;
                    # scan rightward from that column to find most recent non-null value
                    if val is None or not isinstance(val, (int, float)):
                        for scan_col in range(col_idx, min(col_idx + 60, ws.max_column + 1)):
                            candidate = ws.cell(row=row_idx, column=scan_col).value
                            if isinstance(candidate, (int, float)):
                                val = candidate
                                break
                    trust_data[metric_key] = round(float(val), 1) if isinstance(val, (int, float)) else None
                except Exception:
                    trust_data[metric_key] = None

            # Derive corridor_total
            ed   = trust_data.get("corridor_ed")
            ward = trust_data.get("corridor_ward")
            if ed is not None and ward is not None:
                trust_data["corridor_total"] = ed + ward
            elif ed is not None:
                trust_data["corridor_total"] = ed
            elif ward is not None:
                trust_data["corridor_total"] = ward
            else:
                trust_data["corridor_total"] = None

            results[code] = trust_data
            print(f"  {code}: {trust_data}")

        if results:
            break  # Found the data sheet

    return results


def get_last_published_week(wb):
    """
    Try to extract the week-ending date from the workbook.
    Returns ISO date string or None.
    """
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=1, max_row=6, values_only=True):
            for cell in row:
                if isinstance(cell, datetime):
                    return cell.date().isoformat()
                if isinstance(cell, str):
                    m = re.search(r'(\d{1,2})[/\-](\d{1,2})[/\-](20\d{2})', cell)
                    if m:
                        try:
                            d = datetime.strptime(m.group(0).replace("-", "/"), "%d/%m/%Y")
                            return d.date().isoformat()
                        except ValueError:
                            pass
    return None


def load_last_json():
    """Load previously committed JSON for history preservation."""
    try:
        r = requests.get(RAW_URL, timeout=10)
        if r.status_code == 200:
            return r.json()
    except Exception:
        pass
    return {}


def commit_json(content_dict, filepath, message):
    """Commit JSON file to GitHub via API."""
    if not GITHUB_TOKEN:
        print(f"  [DRY RUN — no token] Would commit {filepath}")
        return True

    api_url = f"https://api.github.com/repos/{GITHUB_REPO}/contents/{filepath}"
    headers = {
        "Authorization": f"token {GITHUB_TOKEN}",
        "Accept": "application/vnd.github.v3+json",
    }
    b64 = base64.b64encode(json.dumps(content_dict, indent=2).encode()).decode()
    r = requests.get(api_url, headers=headers)
    sha = r.json().get("sha") if r.status_code == 200 else None
    payload = {"message": message, "content": b64, "branch": "main"}
    if sha:
        payload["sha"] = sha
    r = requests.put(api_url, headers=headers, json=payload)
    if r.status_code in (200, 201):
        print(f"  ✓ Committed {filepath}")
        return True
    print(f"  ✗ Failed {filepath}: {r.status_code} — {r.json().get('message','')}")
    return False


# ── MAIN ──────────────────────────────────────────────────────────────

def main():
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    print(f"\n── Assistiv Corridor Care Fetcher ── {today} ──\n")

    # 1. Find and download the SitRep Excel
    try:
        excel_url = find_sitrep_excel_url()
    except Exception as e:
        print(f"ERROR finding SitRep URL: {e}")
        raise

    excel_bytes = download_excel(excel_url)
    wb = openpyxl.load_workbook(excel_bytes, read_only=True, data_only=True)
    print(f"\nWorkbook sheets: {wb.sheetnames}")

    week_ending = get_last_published_week(wb)
    print(f"Week ending detected: {week_ending or 'unknown'}")

    # 2. Extract Kent trust metrics
    trust_metrics = extract_trust_metrics(wb, KENT_TRUSTS)

    if not trust_metrics:
        print("\nWARNING: No trust data extracted — corridor care fields may not yet be")
        print("present in the SitRep Excel. The new fields were mandated from 6 March 2026")
        print("and publication was committed from May 2026. Check sheet structure manually.")

    # 3. Load history and append this week
    last = load_last_json()
    history = last.get("history", [])

    # Add today's snapshot to history (keep last 52 weeks)
    snapshot = {
        "week_ending": week_ending or today,
        "fetched": today,
        "source_url": excel_url,
        "trusts": {},
    }
    for code, metrics in trust_metrics.items():
        snapshot["trusts"][code] = {
            **metrics,
            "name": KENT_TRUSTS[code]["name"],
            "short": KENT_TRUSTS[code]["short"],
        }

    # Deduplicate by week_ending
    history = [h for h in history if h.get("week_ending") != snapshot["week_ending"]]
    history.append(snapshot)
    history = sorted(history, key=lambda x: x["week_ending"])[-52:]

    # 4. Build current-state trust objects with trend
    trusts_current = {}
    for code, info in KENT_TRUSTS.items():
        metrics = trust_metrics.get(code, {})
        # Compute 4-week trend for corridor_total
        recent = [
            h["trusts"].get(code, {}).get("corridor_total")
            for h in history[-5:]
            if h["trusts"].get(code, {}).get("corridor_total") is not None
        ]
        if len(recent) >= 2:
            delta  = recent[-1] - recent[-2]
            trend  = "rising" if delta > 2 else "falling" if delta < -2 else "stable"
        else:
            delta = None
            trend = "unknown"

        trusts_current[code] = {
            "name":         info["name"],
            "short":        info["short"],
            "districts":    info["districts"],
            "corridor_ed":        metrics.get("corridor_ed"),
            "corridor_ward":      metrics.get("corridor_ward"),
            "corridor_total":     metrics.get("corridor_total"),
            "twelve_hour_waits":  metrics.get("twelve_hour"),
            "beds_occupancy_pct": metrics.get("beds_occ"),
            "delayed_discharges": metrics.get("delayed_discharge"),
            "corridor_delta":     delta,
            "corridor_trend":     trend,
            "week_ending":        week_ending or today,
        }

    # 5. Assemble output JSON
    output = {
        "meta": {
            "generated":    datetime.now(timezone.utc).isoformat(),
            "description":  "Kent NHS Trust corridor care and UEC pressure — Assistiv Systems",
            "version":      "1.0",
            "refresh_type": "weekly — NHS England UEC Daily SitRep (Thursdays)",
            "week_ending":  week_ending or today,
            "source_url":   excel_url,
            "source":       "NHS England UEC Daily Situation Reports 2025-26",
            "licence":      "Open Government Licence v3.0",
            "corridor_care_definition": (
                "Patient has experienced corridor care if they spent ≥45 minutes "
                "in a clinically inappropriate area of an ED or general & acute ward. "
                "NHS England definition, March 2026. "
                "ED count: previous 24h midnight-to-midnight. "
                "Ward count: 8am snapshot of occupied corridor beds."
            ),
            "data_currency_note": (
                "SitRep data published weekly by NHS England, typically every Thursday at 09:30. "
                "Data covers the week ending the previous Sunday. "
                "Corridor care fields introduced March 2026."
            ),
            "trust_codes": list(KENT_TRUSTS.keys()),
        },
        "trusts": trusts_current,
        "history": history,
    }

    # 6. Print summary
    print(f"\n── Corridor Care Summary ── Week ending {week_ending or today} ──")
    for code, t in trusts_current.items():
        print(f"  {code} ({t['short']})")
        print(f"    Corridor (ED):   {t['corridor_ed']}")
        print(f"    Corridor (Ward): {t['corridor_ward']}")
        print(f"    Corridor Total:  {t['corridor_total']} ({t['corridor_trend']})")
        print(f"    12hr Waits:      {t['twelve_hour_waits']}")
        print(f"    Beds Occ%:       {t['beds_occupancy_pct']}")

    # 7. Commit
    msg = f"Corridor care refresh — week ending {week_ending or today}"
    print(f"\nCommitting {GITHUB_FILE}...")
    commit_json(output, GITHUB_FILE, msg)
    print(f"\nDone — {today}")


if __name__ == "__main__":
    main()
